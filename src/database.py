# database.py — Работа с Excel-файлами (база данных)
# ====================================================

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path
from config import (
    DATA_DIR, DB_PIPELINE, DB_EMPLOYEES, DB_BATCHES,
    DB_DEFECTS, DB_TIMESHEETS, DB_SALARY,
    BATCH_AWAITING, BATCH_IN_PROGRESS, BATCH_COMPLETED,
    DEFECT_PENDING, DEFECT_APPROVED, DEFECT_REJECTED,
)


# ============================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ============================================================

def _ensure_dir():
    """Создаёт папку data, если её нет."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)


def _style_header(ws, headers):
    """Оформляет заголовки таблицы."""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Автоширина
    for col, header in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max(len(str(header)) + 4, 15)


def _auto_width(ws):
    """Автоширина колонок по содержимому."""
    for col in ws.columns:
        max_length = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_length + 4, 15)


# ============================================================
# ИНИЦИАЛИЗАЦИЯ ВСЕХ EXCEL-ФАЙЛОВ
# ============================================================

def init_all_databases():
    """Создаёт все Excel-файлы с примерами, если их нет."""
    _ensure_dir()

    _init_pipeline()
    _init_employees()
    _init_batches()
    _init_defects()
    _init_timesheets()
    _init_salary()


def _init_pipeline():
    """Файл: конвейер_настройки.xlsx — настройка конвейера (NO-CODE)"""
    if DB_PIPELINE.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Конвейер"

    headers = ["ID", "Название этапа", "Ответственный (роль)", "Норма времени (мин)", "Требует фотоотчёт"]
    _style_header(ws, headers)

    data = [
        [1, "Раскрой ткани", "закройщик", 45, "Да"],
        [2, "Пошив верха", "швея", 90, "Нет"],
        [3, "Контроль качества", "технолог", 15, "Да"],
        [4, "Упаковка", "упаковщик", 20, "Да"],
    ]

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    _auto_width(ws)
    wb.save(DB_PIPELINE)


def _init_employees():
    """Файл: сотрудники.xlsx — список сотрудников"""
    if DB_EMPLOYEES.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Сотрудники"

    headers = ["ID", "Telegram ID", "Username", "Имя", "Роль", "Дата регистрации"]
    _style_header(ws, headers)

    # Пустой — заполняется при регистрации
    _auto_width(ws)
    wb.save(DB_EMPLOYEES)


def _init_batches():
    """Файл: партии.xlsx — производственные партии"""
    if DB_BATCHES.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Партии"

    headers = [
        "ID партии", "Название", "Текущий этап (ID)", "Статус",
        "Назначен на (ID сотрудника)", "Создана", "Начата", "Завершена",
        "История"
    ]
    _style_header(ws, headers)

    _auto_width(ws)
    wb.save(DB_BATCHES)


def _init_defects():
    """Файл: брак.xlsx — зафиксированные дефекты"""
    if DB_DEFECTS.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Брак"

    headers = [
        "ID", "Партия ID", "Этап", "Сообщил", "Дата",
        "Фото (file_id)", "Статус", "Решил", "Дата решения"
    ]
    _style_header(ws, headers)

    _auto_width(ws)
    wb.save(DB_DEFECTS)


def _init_timesheets():
    """Файл: табель.xlsx — учёт рабочего времени"""
    if DB_TIMESHEETS.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Табель"

    headers = ["ID", "Сотрудник ID", "Имя", "Дата", "Часы", "Записал", "Дата записи"]
    _style_header(ws, headers)

    _auto_width(ws)
    wb.save(DB_TIMESHEETS)


def _init_salary():
    """Файл: зарплата.xlsx — зарплатные ведомости"""
    if DB_SALARY.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Ведомость"

    headers = [
        "ID", "Сотрудник ID", "Имя", "Username (логин)",
        "Месяц", "Год", "Отработано часов", "Ставка (сом/час)", "Начислено"
    ]
    _style_header(ws, headers)

    _auto_width(ws)
    wb.save(DB_SALARY)


# ============================================================
# ЧТЕНИЕ ИЗ EXCEL
# ============================================================

def _read_all_rows(filepath):
    """Читает все строки из Excel (кроме заголовка). Возвращает список словарей."""
    if not filepath.exists():
        return []

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            rows.append(dict(zip(headers, row)))

    wb.close()
    return rows


def _write_all_rows(filepath, headers, rows):
    """Перезаписывает Excel-файл новыми данными."""
    wb = Workbook()
    ws = wb.active
    ws.title = wb.sheetnames[0]

    _style_header(ws, headers)

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))

    _auto_width(ws)
    wb.save(filepath)


# ============================================================
# API ДЛЯ РАБОТЫ С ДАННЫМИ
# ============================================================

# --- КОНВЕЙЕР ---

def get_pipeline():
    """Возвращает список этапов конвейера."""
    rows = _read_all_rows(DB_PIPELINE)
    return [
        {
            "id": r["ID"],
            "name": r["Название этапа"],
            "role": r["Ответственный (роль)"].lower().strip(),
            "norm_minutes": r["Норма времени (мин)"],
            "requires_photo": str(r.get("Требует фотоотчёт", "Нет")).lower() == "да",
        }
        for r in rows
    ]


# --- СОТРУДНИКИ ---

def get_employees():
    return _read_all_rows(DB_EMPLOYEES)


def find_employee_by_user_id(user_id):
    employees = get_employees()
    for e in employees:
        if e["Telegram ID"] == user_id:
            return e
    return None


def find_employee_by_username(username):
    employees = get_employees()
    for e in employees:
        if e["Username"] == username.replace("@", ""):
            return e
    return None


def add_employee(user_id, username, full_name, role):
    employees = get_employees()
    new_id = len(employees) + 1
    from datetime import datetime
    employees.append({
        "ID": new_id,
        "Telegram ID": user_id,
        "Username": username,
        "Имя": full_name,
        "Роль": role,
        "Дата регистрации": datetime.now().strftime("%Y-%m-%d %H:%M"),
    })
    headers = ["ID", "Telegram ID", "Username", "Имя", "Роль", "Дата регистрации"]
    _write_all_rows(DB_EMPLOYEES, headers, employees)


def update_employee_role(user_id, new_role):
    employees = get_employees()
    for e in employees:
        if e["Telegram ID"] == user_id:
            e["Роль"] = new_role
            break
    headers = ["ID", "Telegram ID", "Username", "Имя", "Роль", "Дата регистрации"]
    _write_all_rows(DB_EMPLOYEES, headers, employees)


# --- ПАРТИИ ---

def get_batches():
    return _read_all_rows(DB_BATCHES)


def add_batch(name, first_process_id):
    batches = get_batches()
    new_id = len(batches) + 1
    from datetime import datetime
    batches.append({
        "ID партии": new_id,
        "Название": name,
        "Текущий этап (ID)": first_process_id,
        "Статус": BATCH_AWAITING,
        "Назначен на (ID сотрудника)": "",
        "Создана": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "Начата": "",
        "Завершена": "",
        "История": "",
    })
    headers = [
        "ID партии", "Название", "Текущий этап (ID)", "Статус",
        "Назначен на (ID сотрудника)", "Создана", "Начата", "Завершена", "История"
    ]
    _write_all_rows(DB_BATCHES, headers, batches)
    return new_id


def update_batch(batch_id, updates):
    batches = get_batches()
    for b in batches:
        if b["ID партии"] == batch_id:
            b.update(updates)
            break
    headers = [
        "ID партии", "Название", "Текущий этап (ID)", "Статус",
        "Назначен на (ID сотрудника)", "Создана", "Начата", "Завершена", "История"
    ]
    _write_all_rows(DB_BATCHES, headers, batches)


# --- БРАК ---

def get_defects():
    return _read_all_rows(DB_DEFECTS)


def add_defect(batch_id, process_name, reported_by, reported_by_id, photo_file_id):
    defects = get_defects()
    new_id = len(defects) + 1
    from datetime import datetime
    defects.append({
        "ID": new_id,
        "Партия ID": batch_id,
        "Этап": process_name,
        "Сообщил": reported_by,
        "Дата": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "Фото (file_id)": photo_file_id,
        "Статус": DEFECT_PENDING,
        "Решил": "",
        "Дата решения": "",
    })
    headers = [
        "ID", "Партия ID", "Этап", "Сообщил", "Дата",
        "Фото (file_id)", "Статус", "Решил", "Дата решения"
    ]
    _write_all_rows(DB_DEFECTS, headers, defects)
    return new_id


def update_defect(defect_id, status, resolved_by):
    defects = get_defects()
    from datetime import datetime
    for d in defects:
        if d["ID"] == defect_id:
            d["Статус"] = status
            d["Решил"] = resolved_by
            d["Дата решения"] = datetime.now().strftime("%Y-%m-%d %H:%M")
            break
    headers = [
        "ID", "Партия ID", "Этап", "Сообщил", "Дата",
        "Фото (file_id)", "Статус", "Решил", "Дата решения"
    ]
    _write_all_rows(DB_DEFECTS, headers, defects)


# --- ТАБЕЛЬ ---

def get_timesheets():
    return _read_all_rows(DB_TIMESHEETS)


def add_timesheet(user_id, full_name, date, hours, recorded_by):
    timesheets = get_timesheets()
    new_id = len(timesheets) + 1
    from datetime import datetime
    timesheets.append({
        "ID": new_id,
        "Сотрудник ID": user_id,
        "Имя": full_name,
        "Дата": date,
        "Часы": hours,
        "Записал": recorded_by,
        "Дата записи": datetime.now().strftime("%Y-%m-%d %H:%M"),
    })
    headers = ["ID", "Сотрудник ID", "Имя", "Дата", "Часы", "Записал", "Дата записи"]
    _write_all_rows(DB_TIMESHEETS, headers, timesheets)


# --- ЗАРПЛАТА ---

def get_salary():
    return _read_all_rows(DB_SALARY)


def get_salary_for_user(user_id):
    salary = get_salary()
    return [s for s in salary if s["Сотрудник ID"] == user_id]